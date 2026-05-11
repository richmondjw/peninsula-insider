#!/usr/bin/env python3
"""
Peninsula Insider — Venue Operational Audit
Reads every venue JSON, HTTP-checks each website, scores risk, and writes an Excel report.

Usage:
  python ops/scripts/venue-audit.py
  python ops/scripts/venue-audit.py --out ops/reports/venue-audit-custom.xlsx

Output: ops/reports/venue-audit/venue-audit-YYYY-MM-DD.xlsx
"""

import argparse
import json
import re
import socket
import sys
import urllib.request
import urllib.error
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found — run: pip install openpyxl")
    sys.exit(1)

REPO_ROOT   = Path(__file__).resolve().parents[2]
VENUES_DIR  = REPO_ROOT / "next" / "src" / "content" / "venues"
REPORT_DIR  = REPO_ROOT / "ops" / "reports" / "venue-audit"

TODAY        = date.today()
STALE_ERROR  = 540   # days — 18 months: HIGH risk
STALE_WARN   = 270   # days — 9 months: MEDIUM risk

UA = "Mozilla/5.0 (compatible; PeninsulaInsiderBot/1.0; +https://peninsulainsider.com.au)"

# ---------------------------------------------------------------------------
# Risk colours (Excel fills)
# ---------------------------------------------------------------------------
FILL_HIGH   = PatternFill("solid", fgColor="FFCCCC")   # red tint
FILL_MEDIUM = PatternFill("solid", fgColor="FFF2CC")   # amber tint
FILL_LOW    = PatternFill("solid", fgColor="D9EAD3")   # green tint
FILL_HEADER = PatternFill("solid", fgColor="1C3A5E")   # PI navy
FILL_GROUP  = PatternFill("solid", fgColor="E8EFF7")   # light blue

FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_RISK_H = Font(bold=True, color="9C0006", name="Calibri", size=10)
FONT_RISK_M = Font(bold=True, color="7D4E00", name="Calibri", size=10)
FONT_RISK_L = Font(bold=True, color="1E5C1E", name="Calibri", size=10)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# HTTP check
# ---------------------------------------------------------------------------
def check_website(url: str | None, timeout: int = 8) -> tuple[str, str]:
    """Return (http_status_string, note)."""
    if not url:
        return "NO WEBSITE", "No website field in venue data"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            code = resp.getcode()
            final_url = resp.geturl()
            if final_url != url and any(
                dead in final_url for dead in ["parked", "godaddy", "404", "domain-for-sale"]
            ):
                return f"{code} (PARKED)", "URL redirects to domain-parking page"
            return str(code), ""
    except urllib.error.HTTPError as e:
        return str(e.code), f"HTTP {e.code}"
    except urllib.error.URLError as e:
        reason = str(e.reason)
        if "timed out" in reason.lower() or isinstance(e.reason, socket.timeout):
            return "TIMEOUT", "Connection timed out"
        if "name or service not known" in reason.lower() or "nodename" in reason.lower():
            return "DNS FAIL", "Domain does not resolve — likely closed/expired"
        return "CONN ERROR", reason[:80]
    except Exception as e:
        return "ERROR", str(e)[:80]


# ---------------------------------------------------------------------------
# Risk scoring
# ---------------------------------------------------------------------------
def score_venue(data: dict, http_status: str, http_note: str) -> tuple[str, list[str]]:
    """Return (risk_level, [reasons])."""
    flags: list[str] = []
    high = 0
    medium = 0

    # 1. Explicit closed status
    if data.get("status") == "closed":
        flags.append("status = closed")
        high += 10

    # 2. Last verified age
    lv = data.get("lastVerified")
    if lv:
        try:
            lv_date = datetime.strptime(str(lv), "%Y-%m-%d").date()
            age_days = (TODAY - lv_date).days
            if age_days > STALE_ERROR:
                flags.append(f"Last verified {age_days}d ago (>{STALE_ERROR}d)")
                high += 3
            elif age_days > STALE_WARN:
                flags.append(f"Last verified {age_days}d ago (>{STALE_WARN}d)")
                medium += 2
        except ValueError:
            flags.append("lastVerified date unparseable")
            medium += 1
    else:
        flags.append("No lastVerified date")
        medium += 2

    # 3. Website HTTP
    if http_status in ("DNS FAIL",):
        flags.append(f"Website: {http_status} — {http_note}")
        high += 5
    elif http_status in ("404", "410", "TIMEOUT", "CONN ERROR", "ERROR"):
        flags.append(f"Website: {http_status}" + (f" — {http_note}" if http_note else ""))
        high += 3
    elif http_status == "NO WEBSITE":
        flags.append("No website URL")
        medium += 2
    elif http_status.startswith("5"):
        flags.append(f"Website server error: {http_status}")
        medium += 2
    elif "PARKED" in http_status:
        flags.append(f"Website appears parked/expired: {http_status}")
        high += 4

    # 4. No phone number
    if not data.get("phone"):
        flags.append("No phone number")
        medium += 1

    # 5. No address
    if not data.get("address"):
        flags.append("No address")
        medium += 1

    # 6. sitemapExclude — staging-only venue
    if data.get("sitemapExclude"):
        flags.append("sitemapExclude = true (not on public sitemap)")
        medium += 1

    # 7. No editorNote or signature
    note = (data.get("editorNote") or "").strip()
    sig  = (data.get("signature") or "").strip()
    if not note and not sig:
        flags.append("No editorNote or signature copy")
        medium += 1

    if high >= 5:
        return "HIGH", flags
    elif high >= 2 or medium >= 4:
        return "HIGH", flags
    elif high >= 1 or medium >= 2:
        return "MEDIUM", flags
    elif medium >= 1 or flags:
        return "MEDIUM", flags
    return "LOW", flags


# ---------------------------------------------------------------------------
# Main audit
# ---------------------------------------------------------------------------
def run_audit() -> list[dict]:
    results = []
    venue_files = sorted(f for f in VENUES_DIR.glob("*.json") if not f.name.startswith("_"))
    total = len(venue_files)
    print(f"Auditing {total} venues...")

    for idx, fp in enumerate(venue_files, 1):
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  [{idx}/{total}] ERROR reading {fp.name}: {e}")
            continue

        slug    = data.get("slug", fp.stem)
        name    = data.get("name", slug)
        website = data.get("website", "")

        print(f"  [{idx}/{total}] {name} — checking {website or '(no website)'}...", end=" ", flush=True)
        http_status, http_note = check_website(website)
        print(http_status)

        risk, flags = score_venue(data, http_status, http_note)

        lv = data.get("lastVerified", "")
        try:
            lv_date = datetime.strptime(str(lv), "%Y-%m-%d").date()
            age_days = (TODAY - lv_date).days
        except Exception:
            age_days = None

        results.append({
            "slug":          slug,
            "name":          name,
            "type":          data.get("type", ""),
            "place":         data.get("place", ""),
            "zone":          data.get("zone", ""),
            "website":       website or "",
            "phone":         data.get("phone", ""),
            "address":       data.get("address", ""),
            "last_verified": str(lv) if lv else "",
            "age_days":      age_days,
            "status":        data.get("status", ""),
            "sitemap_excl":  "Yes" if data.get("sitemapExclude") else "",
            "http_status":   http_status,
            "risk":          risk,
            "flags":         " | ".join(flags) if flags else "None",
            "action":        _suggested_action(risk, flags, data),
        })

    results.sort(key=lambda r: (("HIGH", "MEDIUM", "LOW").index(r["risk"]), r["name"]))
    return results


def _suggested_action(risk: str, flags: list[str], data: dict) -> str:
    if data.get("status") == "closed":
        return "Remove from site"
    if any("DNS FAIL" in f or "PARKED" in f for f in flags):
        return "Verify closure — website dead"
    if any("DNS FAIL" in f or "TIMEOUT" in f or "404" in f for f in flags):
        return "Re-verify website + call venue"
    if risk == "HIGH":
        return "Verify still operating before next publish"
    if risk == "MEDIUM":
        return "Schedule re-verification"
    return "Monitor"


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Risk",          14),
    ("Name",          32),
    ("Type",          14),
    ("Place",         16),
    ("Zone",          14),
    ("Website",       38),
    ("HTTP Status",   14),
    ("Phone",         18),
    ("Last Verified", 14),
    ("Age (days)",    11),
    ("Status",        10),
    ("Sitemap Excl.", 11),
    ("Risk Flags",    55),
    ("Suggested Action", 34),
]


def write_excel(results: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()

    # -----------
    # Summary tab
    # -----------
    ws_sum = wb.active
    ws_sum.title = "Summary"

    high   = [r for r in results if r["risk"] == "HIGH"]
    medium = [r for r in results if r["risk"] == "MEDIUM"]
    low    = [r for r in results if r["risk"] == "LOW"]

    summary_rows = [
        ["Peninsula Insider — Venue Operational Audit"],
        [f"Generated: {TODAY.isoformat()}   |   Total venues: {len(results)}"],
        [],
        ["Risk Level", "Count", "Description"],
        ["HIGH",   len(high),   "Likely closed or website dead — verify immediately"],
        ["MEDIUM", len(medium), "Stale data or minor gaps — schedule re-verification"],
        ["LOW",    len(low),    "Recently verified, website live"],
    ]
    for row in summary_rows:
        ws_sum.append(row)

    ws_sum["A1"].font = Font(bold=True, size=14, name="Calibri", color="1C3A5E")
    for r, fill, font in [
        (5, FILL_HIGH,   FONT_RISK_H),
        (6, FILL_MEDIUM, FONT_RISK_M),
        (7, FILL_LOW,    FONT_RISK_L),
    ]:
        for c in range(1, 4):
            ws_sum.cell(row=r, column=c).fill = fill
        ws_sum.cell(row=r, column=1).font = font

    ws_sum.column_dimensions["A"].width = 14
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 55

    # Write legend note
    ws_sum.append([])
    ws_sum.append(["Notes"])
    ws_sum.append(["• HIGH = status:closed, dead/missing website, or last verified >18 months ago"])
    ws_sum.append(["• MEDIUM = last verified 9–18 months, or website errors, or missing data"])
    ws_sum.append(["• LOW = website live 200 OK and verified within 9 months"])
    ws_sum.append(["• 'HTTP Status' is a live check run at audit time — a 200 means the domain resolves and returns a page"])
    ws_sum.append(["• DNS FAIL = domain does not exist — strong signal of closure"])
    ws_sum.append(["• TIMEOUT = server not responding — could be maintenance, could be gone"])

    # ----------------
    # Detail tabs (one per risk level, then all)
    # ----------------
    for label, subset in [("HIGH Risk", high), ("MEDIUM Risk", medium), ("LOW Risk", low), ("All Venues", results)]:
        ws = wb.create_sheet(label)
        _write_detail_sheet(ws, subset)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _write_detail_sheet(ws, results: list[dict]) -> None:
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)

    # Style header row
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    field_order = [
        "risk", "name", "type", "place", "zone", "website", "http_status",
        "phone", "last_verified", "age_days", "status", "sitemap_excl", "flags", "action",
    ]

    for row_data in results:
        row = [row_data.get(f, "") for f in field_order]
        ws.append(row)
        row_idx = ws.max_row

        risk = row_data["risk"]
        fill = FILL_HIGH if risk == "HIGH" else FILL_MEDIUM if risk == "MEDIUM" else FILL_LOW
        risk_font = FONT_RISK_H if risk == "HIGH" else FONT_RISK_M if risk == "MEDIUM" else FONT_RISK_L

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (13, 14)))

        # Colour the Risk cell
        risk_cell = ws.cell(row=row_idx, column=1)
        risk_cell.fill = fill
        risk_cell.font = risk_font
        risk_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Website as hyperlink
        website = row_data.get("website", "")
        if website:
            ws_cell = ws.cell(row=row_idx, column=6)
            ws_cell.hyperlink = website
            ws_cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)

        # Row height
        flags_len = len(row_data.get("flags", ""))
        ws.row_dimensions[row_idx].height = max(18, min(72, flags_len // 3))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=str, default=None)
    args = parser.parse_args()

    results = run_audit()

    out_path = Path(args.out) if args.out else REPORT_DIR / f"venue-audit-{TODAY.isoformat()}.xlsx"
    write_excel(results, out_path)

    high = sum(1 for r in results if r["risk"] == "HIGH")
    med  = sum(1 for r in results if r["risk"] == "MEDIUM")
    low  = sum(1 for r in results if r["risk"] == "LOW")

    print(f"\nAudit complete: {len(results)} venues")
    print(f"  HIGH:   {high}")
    print(f"  MEDIUM: {med}")
    print(f"  LOW:    {low}")
    print(f"\nReport: {out_path}")


if __name__ == "__main__":
    main()
