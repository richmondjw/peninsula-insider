#!/usr/bin/env python3
"""
import-events.py — turn the master events spreadsheet into the events
content collection, preserving editorial overlay across re-imports.

Usage:
    python scripts/import-events.py [--xlsx PATH] [--out DIR] [--dry-run]

Defaults:
    --xlsx     C:\\Users\\James\\Downloads\\mornington_peninsula_events_calendar_database.xlsx
    --out      next/src/content/events/
    --dry-run  Print the diff but don't write anything

The script reads the manifest from import-events.config.mjs (just the
field lists — JS module is parsed lightly, no Node required) so the
machine-vs-editorial ownership stays in one place.

Behaviour:
    - For each row in the "Events Database" sheet, generate a JSON file.
    - If a JSON file with the same slug already exists, the editorial
      fields are preserved. Machine + derived fields are overwritten.
    - The "Excluded Review" sheet is ignored (historical context only).
    - Reports written to ops/reports/events/import-YYYY-MM-DD.md.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─── Manifest (mirrors import-events.config.mjs) ─────────────────────────────
# Kept inline rather than parsing the .mjs to keep the script dependency-free.
# If you change one, change the other; cron job 13 (overlay-protection-audit)
# will catch divergence.

MACHINE_OWNED_FIELDS = {
    'eventId', 'title', 'summary', 'description',
    'startDate', 'endDate', 'startTime', 'endTime', 'season', 'month',
    'venueName', 'venueRegion', 'suburb', 'streetAddress', 'coordinates',
    'indoorOutdoor',
    'bookingUrl', 'ticketingUrl', 'officialEventUrl', 'bookingRequired',
    'freePaid', 'priceRange',
    'subcategory', 'recurrence', 'recurrenceNote',
    'suitableFor', 'familyFriendly', 'petFriendly', 'accessibilityNotes',
    'weatherDependency',
    'organiser',
    'primarySourceUrl', 'secondarySourceUrl', 'verificationStatus',
    'lastCheckedDate', 'visitorAppealScore', 'editorialPriority',
    'nearbyAttractions', 'suggestedItineraryPairing',
    'internalNotes', 'manualFollowUpRequired',
}

DERIVED_FIELDS = {
    'priceTier', 'weatherShape', 'audienceTags', 'kidsGradeAuto',
    'category', 'place', 'venue', 'nearestVenues', 'nextOccurrence',
}

EDITORIAL_OWNED_FIELDS = {
    'kidsGrade', 'kidsGradeNote',
    'worthTheDrive', 'firstTimer',
    'skipThis', 'skipReason', 'skipInstead',
    'editorVerdict', 'whyWeCare', 'standoutOfMonth', 'pairingProse',
    'editorVisited', 'featuredInDispatch', 'relatedArticles',
    'lens', 'editorNote', 'heroImage',
}

LIFECYCLE_FIELDS = {'slug', 'status', 'publishedAt', 'sitemapExclude'}

# Subcategory + category -> PI category enum (first match wins)
CATEGORY_MAP = [
    (re.compile(r'cellar door', re.I), 'cellar-door'),
    (re.compile(r'exhibition|gallery', re.I), 'exhibition'),
    (re.compile(r'market', re.I), 'market'),
    (re.compile(r'winery|wine', re.I), 'food-wine'),
    (re.compile(r'food|restaurant|dining', re.I), 'food-wine'),
    (re.compile(r'music|concert|gig', re.I), 'live-music'),
    (re.compile(r'race|sport', re.I), 'racing-sport'),
    (re.compile(r'family|kids|children', re.I), 'family-programs'),
    (re.compile(r'walk|hike|nature|outdoor', re.I), 'nature'),
    (re.compile(r'spa|wellness|hot springs', re.I), 'wellness'),
    (re.compile(r'writers|talks|ideas|literary', re.I), 'writers-ideas'),
    (re.compile(r'civic|council|community', re.I), 'community'),
    (re.compile(r'festival|major event', re.I), 'festival'),
    (re.compile(r'art|culture', re.I), 'arts'),
]

SUBURB_TO_PLACE = {
    'mornington': 'mornington',
    'mount martha': 'mount-martha',
    'mt martha': 'mount-martha',
    'red hill': 'red-hill',
    'red hill south': 'red-hill',
    'sorrento': 'sorrento',
    'portsea': 'portsea',
    'blairgowrie': 'blairgowrie',
    'rye': 'rye',
    'rosebud': 'rosebud',
    'dromana': 'dromana',
    'safety beach': 'safety-beach',
    'main ridge': 'main-ridge',
    'cape schanck': 'cape-schanck',
    'flinders': 'flinders',
    'shoreham': 'shoreham',
    'balnarring': 'balnarring',
    'merricks': 'merricks',
    'merricks north': 'merricks',
    'hastings': 'hastings',
    'tuerong': 'tuerong',
    'moorooduc': 'moorooduc',
    'point nepean': 'point-nepean',
}


# ─── Helpers ─────────────────────────────────────────────────────────────────


def slugify(name: str) -> str:
    """Lowercase, kebab-case, ASCII-only, max 80 chars."""
    s = name.lower()
    # Replace common encoding artefacts the spreadsheet ships with
    s = s.replace('—', '-').replace('–', '-').replace('�', '')
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip('-')
    return s[:80]


def clean(s: Any) -> Any:
    """Strip whitespace, normalise encoding artefacts, return None for empties."""
    if s is None:
        return None
    if isinstance(s, str):
        s = s.replace('�', '').strip()
        if not s:
            return None
    return s


def to_iso_date(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def derive_category(subcategory: str | None, primary_category: str | None) -> str:
    """Subcategory wins because it's more specific."""
    haystack = ' '.join(filter(None, [subcategory, primary_category]))
    for pattern, category in CATEGORY_MAP:
        if pattern.search(haystack):
            return category
    return 'community'  # safe fallback; surfaces in editorial review


def derive_place(suburb: str | None) -> str | None:
    if not suburb:
        return None
    return SUBURB_TO_PLACE.get(suburb.strip().lower())


def derive_price_tier(free_paid: str | None, price_range: str | None) -> str:
    if not free_paid:
        return 'unknown'
    fp_lower = free_paid.lower()
    if 'free' in fp_lower and 'paid' not in fp_lower:
        return 'free'
    # Pull a number from the price range string if present
    if price_range:
        nums = re.findall(r'(\d+(?:\.\d+)?)', price_range)
        if nums:
            high = max(float(n) for n in nums)
            if high < 50:
                return 'under-50'
            if high < 150:
                return '50-150'
            return 'over-150'
    return 'unknown'


def derive_weather_shape(indoor_outdoor: str | None, weather_dep: str | None) -> str:
    io_lower = (indoor_outdoor or '').lower()
    wd_lower = (weather_dep or '').lower()
    if 'indoor' in io_lower and 'outdoor' not in io_lower:
        return 'all-weather'
    if 'weather independent' in wd_lower or 'none' in wd_lower:
        return 'all-weather'
    if 'minimal' in wd_lower or 'low' in wd_lower:
        return 'wet-friendly'
    if 'high' in wd_lower:
        return 'fair-weather-only'
    if 'outdoor' in io_lower:
        return 'fair-weather-only'
    return 'unknown'


def derive_audience_tags(suitable_for: str | None,
                         family_friendly: str | None,
                         pet_friendly: str | None) -> list[str]:
    tags: set[str] = set()
    text = (suitable_for or '').lower()
    if 'couple' in text:
        tags.add('couples')
    if 'famil' in text or 'kids' in text or 'children' in text:
        tags.add('families')
    if 'solo' in text:
        tags.add('solo')
    if 'group' in text:
        tags.add('groups')
    if 'first-time' in text or 'first time' in text or 'visitor' in text:
        tags.add('first-timers')
    if 'local' in text:
        tags.add('locals')
    if 'cultural' in text or 'art' in text:
        tags.add('cultural-visitors')
    if 'food' in text or 'foodie' in text:
        tags.add('foodies')
    if 'all ages' in text or 'all-ages' in text:
        tags.add('all-ages')
    if 'adult' in text and 'famil' not in text:
        tags.add('adults-only')
    if 'music' in text:
        tags.add('music-fans')
    if 'art lover' in text:
        tags.add('art-lovers')
    # Family flag override
    if isinstance(family_friendly, str) and family_friendly.lower().startswith('y'):
        tags.add('families')
    return sorted(tags)


def derive_lens_auto(visitor_appeal: int | None,
                     weather_shape: str,
                     kids_grade_auto: str | None,
                     price_tier: str,
                     booking_required: str | None,
                     ticketing_url: str | None,
                     visitor_appeal_threshold_pick: int = 4) -> list[str]:
    """Auto-derived lens tags. Editor's `lens` array still wins on re-import.
    These tags fill the editorial browse rows on the hub for the bulk of
    machine-imported events that the editor hasn't touched yet."""
    tags: list[str] = []
    appeal = visitor_appeal or 0
    if appeal >= visitor_appeal_threshold_pick:
        tags.append('weekend-pick')
    if appeal >= 5:
        tags.append('worth-the-drive')
    if weather_shape == 'all-weather':
        tags.append('rainy-day')
    if kids_grade_auto in ('A', 'B'):
        tags.append('family-saturday')
    if price_tier == 'free':
        tags.append('free')
    booking_lower = (booking_required or '').lower()
    if 'no' in booking_lower or 'walk' in booking_lower:
        tags.append('walk-in')
    elif 'yes' in booking_lower and ticketing_url:
        tags.append('ticketed')
    return tags


def derive_kids_grade_auto(family_friendly: str | None,
                           accessibility: str | None,
                           weather_shape: str) -> str | None:
    """Conservative auto-grade. Editor's kidsGrade always wins."""
    if not isinstance(family_friendly, str):
        return None
    ff = family_friendly.lower()
    if not ff.startswith('y'):
        return 'not-for-kids'
    # Family-friendly = Yes; differentiate by weather + accessibility
    if weather_shape == 'all-weather':
        return 'B'
    return 'C'  # outdoor + family-friendly is C until editor reviews


def normalise_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower().startswith('y')
    return False


def parse_organiser(name: Any, website: Any, contact: Any,
                    instagram: Any, facebook: Any) -> dict | None:
    fields = {
        'name': clean(name),
        'website': clean(website),
        'contact': clean(contact),
        'instagram': clean(instagram),
        'facebook': clean(facebook),
    }
    fields = {k: v for k, v in fields.items() if v}
    return fields if fields else None


def parse_url(v: Any) -> str | None:
    """Spreadsheet has multi-URL strings ('a | b') in some cells; take the first."""
    s = clean(v)
    if not s:
        return None
    first = s.split('|')[0].strip()
    return first if first.startswith('http') else None


def parse_int(v: Any) -> int | None:
    if v is None or v == '':
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def parse_coord(v: Any) -> float | None:
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ─── Row → event JSON ────────────────────────────────────────────────────────


@dataclass
class ImportStats:
    created: list[str] = field(default_factory=list)
    updated: list[str] = field(default_factory=list)
    preserved_overlay: list[str] = field(default_factory=list)
    no_place_match: list[str] = field(default_factory=list)
    skipped: list[tuple[str, str]] = field(default_factory=list)


def row_to_event(row: dict, existing: dict | None) -> dict | None:
    """Build an event dict from a spreadsheet row.

    `existing` is the parsed JSON from disk if a file with this slug
    already exists. Editorial fields from `existing` are preserved.
    """
    title = clean(row.get('Event Name'))
    if not title:
        return None
    start_date = to_iso_date(row.get('Start Date'))
    if not start_date:
        return None  # without a start date the event is unschedulable

    slug = slugify(title)
    suburb = clean(row.get('Suburb'))
    primary_category = clean(row.get('Primary Category'))
    subcategory = clean(row.get('Subcategory'))
    indoor_outdoor = clean(row.get('Indoor / Outdoor'))
    weather_dependency = clean(row.get('Weather Dependency'))
    free_paid = clean(row.get('Free / Paid'))
    price_range = clean(row.get('Price Range'))

    # Coordinates (only if both lat and lng parse cleanly)
    lat = parse_coord(row.get('Latitude'))
    lng = parse_coord(row.get('Longitude'))
    coordinates = {'lat': lat, 'lng': lng} if (lat is not None and lng is not None) else None

    organiser = parse_organiser(
        row.get('Organiser Name'),
        row.get('Organiser Website'),
        row.get('Organiser Contact'),
        row.get('Instagram URL'),
        row.get('Facebook URL'),
    )

    # Machine + derived fields
    event = {
        'slug': slug,
        'eventId': clean(row.get('Event ID')),
        'title': title,
        'summary': (clean(row.get('Description')) or title)[:300],
        'description': clean(row.get('Description')),
        'startDate': start_date,
        'endDate': to_iso_date(row.get('End Date')),
        'startTime': clean(row.get('Start Time')),
        'endTime': clean(row.get('End Time')),
        'season': (clean(row.get('Season')) or '').lower() or None,
        'month': clean(row.get('Month')),
        'venueName': clean(row.get('Venue Name')),
        'venueRegion': clean(row.get('Region / Area')),
        'suburb': suburb,
        'streetAddress': clean(row.get('Street Address')),
        'coordinates': coordinates,
        'indoorOutdoor': indoor_outdoor,
        'bookingUrl': parse_url(row.get('Ticketing URL')),
        'ticketingUrl': parse_url(row.get('Ticketing URL')),
        'officialEventUrl': clean(row.get('Official Event URL')),
        'bookingRequired': clean(row.get('Booking Required')),
        'freePaid': free_paid,
        'priceRange': price_range,
        'priceTier': derive_price_tier(free_paid, price_range),
        'category': derive_category(subcategory, primary_category),
        'subcategory': subcategory,
        'recurrence': normalise_recurrence(row.get('Recurrence')),
        'recurrenceNote': clean(row.get('Recurrence')),
        'suitableFor': clean(row.get('Suitable For')),
        'audienceTags': derive_audience_tags(
            clean(row.get('Suitable For')),
            clean(row.get('Family Friendly')),
            clean(row.get('Pet Friendly')),
        ),
        'familyFriendly': normalise_bool(row.get('Family Friendly')),
        'petFriendly': normalise_bool(row.get('Pet Friendly')),
        'accessibilityNotes': clean(row.get('Accessibility Notes')),
        'weather': 'mixed',  # legacy field, default
        'weatherDependency': weather_dependency,
        'weatherShape': derive_weather_shape(indoor_outdoor, weather_dependency),
        'organiser': organiser,
        'primarySourceUrl': clean(row.get('Primary Source URL')),
        'secondarySourceUrl': clean(row.get('Secondary Source URL')),
        'verificationStatus': clean(row.get('Verification Status')),
        'lastCheckedDate': to_iso_date(row.get('Last Checked Date')),
        'visitorAppealScore': parse_int(row.get('Visitor Appeal Score')),
        'editorialPriority': parse_int(row.get('Commercial / Editorial Priority')),
        'nearbyAttractions': clean(row.get('Nearby Attractions')),
        'suggestedItineraryPairing': clean(row.get('Suggested Itinerary Pairing')),
        'internalNotes': clean(row.get('Issues / Notes')),
        'manualFollowUpRequired': normalise_bool(row.get('Manual Follow-up Required')),
        'kidsGradeAuto': derive_kids_grade_auto(
            clean(row.get('Family Friendly')),
            clean(row.get('Accessibility Notes')),
            derive_weather_shape(indoor_outdoor, weather_dependency),
        ),
        'status': 'published',
        'publishedAt': start_date,  # initial publish, editor can override
    }

    # Auto-derive worthTheDrive when the data clearly says so — keeps the
    # editor field as the override (editorial-owned), but populates a
    # sensible default from visitor appeal score 5 + verified status.
    visitor_appeal_int = parse_int(row.get('Visitor Appeal Score')) or 0
    verification = (clean(row.get('Verification Status')) or '').lower()
    is_verified = 'verified' in verification or 'recurring, next date confirmed' in verification
    if visitor_appeal_int >= 5 and is_verified:
        event['worthTheDrive'] = True

    # Auto-derive lens tags so the hub's editorial browse rows fill out
    # for machine-imported events. Editor's lens array (preserved as
    # editorial overlay) still wins on re-import.
    event['lens'] = derive_lens_auto(
        visitor_appeal_int,
        event.get('weatherShape', 'unknown'),
        event.get('kidsGradeAuto'),
        event.get('priceTier', 'unknown'),
        event.get('bookingRequired'),
        event.get('ticketingUrl'),
    )

    # Place ref: only set if the suburb maps cleanly. Otherwise leave unset
    # so the editor can map it manually (showing in needs-review queue).
    place = derive_place(suburb)
    if place:
        event['place'] = place

    # Drop None/empty values to keep JSON clean
    event = {k: v for k, v in event.items() if v not in (None, '', [])}

    # ─── Preserve editorial overlay from existing file ───────────────────
    if existing:
        for f in EDITORIAL_OWNED_FIELDS:
            if f in existing and existing[f] not in (None, '', [], False):
                event[f] = existing[f]
        # Lifecycle: editor can override slug/status/publishedAt/sitemapExclude
        for f in LIFECYCLE_FIELDS:
            if f in existing and f not in ('slug',):  # slug always derived from title
                event[f] = existing[f]

    return event


def normalise_recurrence(v: Any) -> str:
    """Map spreadsheet's free-text recurrence to the schema enum."""
    s = (clean(v) or '').lower()
    if not s:
        return 'one-off'
    if 'weekly' in s:
        return 'weekly'
    if 'monthly' in s:
        return 'monthly'
    if 'annual' in s:
        return 'annual'
    if 'seasonal' in s:
        return 'seasonal'
    if 'ongoing' in s or 'recurring' in s:
        return 'ongoing'
    return 'one-off'


# ─── Main ────────────────────────────────────────────────────────────────────


def import_events(xlsx_path: Path, out_dir: Path, dry_run: bool = False) -> ImportStats:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Events Database']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    stats = ImportStats()

    out_dir.mkdir(parents=True, exist_ok=True)

    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if not any(row_cells):
            continue
        row = dict(zip(headers, row_cells))
        title = clean(row.get('Event Name'))
        if not title:
            continue

        slug = slugify(title)
        out_path = out_dir / f"{slug}.json"
        existing = None
        if out_path.exists():
            try:
                existing = json.loads(out_path.read_text(encoding='utf-8'))
            except Exception as e:
                stats.skipped.append((slug, f"failed to parse existing: {e}"))
                continue

        event = row_to_event(row, existing)
        if not event:
            stats.skipped.append((slug, 'no title or start date'))
            continue

        had_overlay = existing and any(
            f in existing for f in EDITORIAL_OWNED_FIELDS
            if existing.get(f) not in (None, '', [], False)
        )
        if had_overlay:
            stats.preserved_overlay.append(slug)

        if not event.get('place'):
            stats.no_place_match.append(f"{slug} (suburb: {event.get('suburb')})")

        if dry_run:
            stats.created.append(slug) if not existing else stats.updated.append(slug)
            continue

        if existing:
            stats.updated.append(slug)
        else:
            stats.created.append(slug)
            event['publishedAt'] = datetime.now().date().isoformat()

        out_path.write_text(json.dumps(event, indent=2, ensure_ascii=False) + '\n',
                            encoding='utf-8')

    return stats


def write_report(stats: ImportStats, report_dir: Path) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().date().isoformat()
    report_path = report_dir / f"import-{today}.md"
    lines = [
        f"# Events import report — {today}",
        '',
        f"- Created: {len(stats.created)}",
        f"- Updated: {len(stats.updated)}",
        f"- Editorial overlay preserved: {len(stats.preserved_overlay)}",
        f"- No place match (needs editor review): {len(stats.no_place_match)}",
        f"- Skipped: {len(stats.skipped)}",
        '',
    ]
    if stats.no_place_match:
        lines.append('## Events without a place ref')
        lines.append('')
        for entry in stats.no_place_match:
            lines.append(f"- {entry}")
        lines.append('')
    if stats.skipped:
        lines.append('## Skipped rows')
        lines.append('')
        for slug, reason in stats.skipped:
            lines.append(f"- {slug}: {reason}")
        lines.append('')
    if stats.preserved_overlay:
        lines.append('## Events with preserved editorial overlay')
        lines.append('')
        for slug in stats.preserved_overlay:
            lines.append(f"- {slug}")
        lines.append('')
    report_path.write_text('\n'.join(lines), encoding='utf-8')
    return report_path


def main() -> int:
    here = Path(__file__).resolve().parent
    repo_root = here.parent.parent  # next/scripts/ → repo root
    default_xlsx = Path(r'C:\Users\James\Downloads\mornington_peninsula_events_calendar_database.xlsx')
    default_out = here.parent / 'src' / 'content' / 'events'
    default_report = repo_root / 'ops' / 'reports' / 'events'

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--xlsx', type=Path, default=default_xlsx)
    parser.add_argument('--out', type=Path, default=default_out)
    parser.add_argument('--report-dir', type=Path, default=default_report)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: {args.xlsx} not found", file=sys.stderr)
        return 1

    print(f"Reading: {args.xlsx}")
    print(f"Writing to: {args.out}")
    if args.dry_run:
        print("(dry run — no files will be written)")
    print()

    stats = import_events(args.xlsx, args.out, dry_run=args.dry_run)

    print(f"Created: {len(stats.created)}")
    print(f"Updated: {len(stats.updated)}")
    print(f"Editorial overlay preserved: {len(stats.preserved_overlay)}")
    print(f"No place match: {len(stats.no_place_match)}")
    print(f"Skipped: {len(stats.skipped)}")

    if not args.dry_run:
        report_path = write_report(stats, args.report_dir)
        print(f"\nReport: {report_path}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
